#!/usr/bin/python3
import configparser
import csv
import git
import depot_tools
import os
import requests
import shutil
import time

from depot_tools.gerrit_util import CreateHttpConn
from depot_tools.gerrit_util import GerritError
from depot_tools.gerrit_util import ReadHttpJsonResponse

from openpyxl import Workbook

class BaseCrawler:
	def __init__(self, cfg_path):
		self.__users = []
		self.__initialized = False

		self.__config = configparser.ConfigParser()

		try:
			self.__config.read(cfg_path)
		except configparser.Error:
			print('fail to parse config file %s' % (cfg_path))
			return

		for key in self.__config:
			if key.split(' ')[0] not in ['user', 'gerrit', 'git', 'github', 'patchwork', 'DEFAULT']:
				print('invalid section name "%s"' % (key))
				return

		for key in self.__config:
			if key.split(' ')[0] == 'user':
				if self.__config[key]['disable'].lower() != 'false':
					continue

				self.__users.append({'name': self.__config[key]['name'],
						     'emails': [self.__config[key]['email1'], self.__config[key]['email2']],
						     'function': self.__config[key]['function'],
						     'github username': self.__config[key]['github username'],
						     })

		self.__initialized = True

		return

	def get_initialized(self):
		return self.__initialized

	def get_config(self):
		return self.__config

	def get_users(self):
		return self.__users

	def get_user(self, github_username = '', email = ''):
		if self.__initialized == False:
			return None

		for user in self.__users:
			if github_username != '' and user['github username'] == github_username:
				return user
			if email != '' and email in user['emails']:
				return user

		return None

	def export_csv_file(self, report_directory, report_name, csv_fields, rows):
		if self.__initialized == False:
			return False

		csv_path = '%s/%s.csv' % (report_directory, report_name)
		print('export data to %s' % (csv_path))

		with open(csv_path, 'w', newline = '') as csv_file:
			csv_writer = csv.DictWriter(csv_file, fieldnames = csv_fields)
			csv_writer.writeheader()

			for row in rows:
				csv_writer.writerow(row)

		print('- %d row(s) saved' % (len(rows)))

		return True

	def export_excel_file(self, report_directory, report_name, csv_fields, date_field, rows):
		if self.__initialized == False:
			return False

		excel_path = '%s/%s.xlsx' % (report_directory, report_name)
		print('export data to %s' % (excel_path))

		book = Workbook()
		sheet = book.active

		# add one sheet for counts of each user
		sheet.title = 'summary'
		counts = {}
		for user in self.__users:
			counts[user['name']] = []

		years = []

		for row in rows:
			year = row[date_field].split('-')[0]

			if year not in years:
				for user in self.__users:
					counts[user['name']].append(0)
				years.append('%s' % (year))

			idx = years.index(year)

			counts[row['user_name']][idx] += 1

		now = time.localtime()
		timestamp = time.strftime('%Y-%m%d', now)

		sheet.append(['summary of %s (%s)' % (report_name, timestamp)])

		data = ['']
		for year in years:
			data.append(year)
		sheet.append(data)

		for user in self.__users:
			data = [user['name']]
			for count in counts[user['name']]:
				data.append('%s' % (count))
			sheet.append(data)

		# add one sheet for all data
		sheet = book.create_sheet('all')
		sheet.append(csv_fields)

		for row in rows:
			data = []
			for field in csv_fields:
				data.append(row[field])

			sheet.append(data)

		print('- sheet "%s" added' % (sheet.title))

		# add one sheet for each year
		year = ''
		for row in rows:
			this_year = row[date_field].split('-')[0]
			if year != this_year:
				if year != '':
					print('- sheet "%s" added' % (sheet.title))

				year = this_year

				# create new sheet
				sheet = book.create_sheet(year)
				sheet.append(csv_fields)

			data = []
			for field in csv_fields:
				data.append(row[field])

			sheet.append(data)

		if year != '':
			print('- sheet "%s" added' % (sheet.title))

		# add one sheet for each user
		for user in self.__users:
			# create new sheet
			sheet = book.create_sheet(user['name'])
			sheet.append(csv_fields)

			for row in rows:
				if row['user_name'] != user['name']:
					continue

				data = []
				for field in csv_fields:
					data.append(row[field])

				sheet.append(data)

			print('- sheet "%s" added' % (sheet.title))

		book.save(excel_path)

		return True

class GerritCrawler(BaseCrawler):
	__csv_fields = ['user_name', 'user_function', 'repo_name', 'repo_url', 'project', 'branch', 'change_id', 'subject', 'status', 'created', 'updated', 'submitted', 'insertions', 'deletions', 'owner']
	__report_name = 'gerrit-changes'

	def __init__(self, cfg_path):
		self.__servers = []
		self.__initialized = False

		# call parent's init
		super().__init__(cfg_path)

		if super().get_initialized() == False:
			return

		config = super().get_config()

		for key in config:
			if key.split(' ')[0] == 'gerrit':
				if config[key]['disable'].lower() != 'false':
					continue

				self.__servers.append({'name': config[key]['name'],
						       'url': config[key]['url'],
						      })

		self.__initialized = True

		return

	def get_changes(self):
		# gerrit REST API doc:
		# https://gerrit-review.googlesource.com/Documentation/rest-api.html

		def useDateTime(element):
			# 'created': '2021-11-02 07:16:18.000000000'
			return element['created']

		self.__changes = []

		if self.__initialized == False:
			return self.__changes

		for server in self.__servers:
			print('query changes from gerrit server "%s"' % (server['name']))

			for user in self.get_users():
				for email in user['emails']:
					print('query changes for "%s <%s>"' % (user['name'], email))

					start = 0

					while True:
						more_changes = False

						try:
							changes = ReadHttpJsonResponse(CreateHttpConn(server['url'], 'changes/?q=owner:' + email + '&start=' + str(start)))
						except GerritError as error:
							print('- gerrit error: %s' % (error.message))
							break

						print('- %d change(s) found' % (len(changes)))

						for change in changes:
							# optional field, and not every merged change has this field set
							if 'submitted' not in change.keys():
								change['submitted'] = ''

							# ChangeInfo
							# https://gerrit-review.googlesource.com/Documentation/rest-api-changes.html#change-info
							self.__changes.append({'user_name': user['name'],
										   'user_function': user['function'],
										   'repo_name': server['name'],
										   'repo_url': server['url'],
										   'project': change['project'],
										   'branch': change['branch'],
										   'change_id': change['change_id'],
										   'subject': change['subject'],
										   'status': change['status'],
										   'created': change['created'],
										   'updated': change['updated'],
										   'submitted': change['submitted'],
										   'insertions': change['insertions'],
										   'deletions': change['deletions'],
										   'owner': email,
										  })

							if '_more_changes' in change.keys() and change['_more_changes'] == True:
								start += len(changes)
								more_changes = True

						if more_changes == False:
							break


		# sort the changes by date
		self.__changes.sort(key = useDateTime)

		return self.__changes

	def export_csv_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_csv_file(report_directory, self.__report_name, self.__csv_fields, self.__changes)

		return ret

	def export_excel_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_excel_file(report_directory, self.__report_name, self.__csv_fields, 'created', self.__changes)

		return ret

class GitCrawler(BaseCrawler):
	__csv_fields = ['user_name', 'user_function', 'commit_hash', 'author_email', 'author_date', 'committer_email', 'committer_date', 'subject', 'status']
	__report_name = 'git-commits'

	def __init__(self, cfg_path):
		self.__repos = []
		self.__initialized = False

		# call parent's init
		super().__init__(cfg_path)

		if super().get_initialized() == False:
			return

		config = super().get_config()

		for key in config:
			if key.split(' ')[0] == 'git':
				if config[key]['disable'].lower() != 'false':
					continue

				self.__repos.append({'name': config[key]['name'],
						     'url': config[key]['url'],
						     'branch': config[key]['branch'],
						    })

		# prepare the parameter for git log command
		self.__log_param = []

		for user in self.get_users():
			for email in user['emails']:
				self.__log_param.append('--author=%s' % (email))

		# %H: commit hash
		# %ae: author email
		# %aI: author date, strict ISO 8601 format
		# %ce: committer email
		# %cI: committer date, strict ISO 8601 format
		# %s: subject
		self.__log_param.append('--pretty=format:%H%x09%ae%x09%aI%x09%ce%x09%cI%x09%s')
		self.__log_param.append('--reverse')

		# create the root directory for repos
		self.__repo_root = os.path.abspath('./repo')

		if os.path.isdir(self.__repo_root) == False:
			os.mkdir(self.__repo_root)

		self.__initialized = True

		return

	def __open_repo(self, repo):
		repo_path = os.path.abspath(self.__repo_root + '/' + repo['name'])

		if os.path.isdir(repo_path) == False:
			# repo directory not exist
			print('- clone git repo from %s' % (repo['url']))
			repository = git.Repo.clone_from(repo['url'], repo_path)
		else:
			print('- open git repo at %s' % (repo_path))
			repository = git.Repo(repo_path)

		if repository.__class__ is git.Repo:
			# check if repo is healthy
			if repository.is_dirty(untracked_files = True):
				print('- warning, repo is dirty')

			if repository.remotes.origin.exists() == False:
				print('- warning, remote origin does not exist')

			return repository

		# repo may be corrupted...
		print('- repo corrupted, delete entire repo')
		try:
			shutil.rmtree(repo_path)
		except OSError as error:
			pass

		return None

	def get_commits(self):
		def useDateTime(element):
			# 'created': '2021-08-10T11:47:55+02:00'
			return element['committer_date']

		self.__commits = []

		if self.__initialized == False:
			return self.__commits

		hash_cache = []

		for repo in self.__repos:
			print('query commits from git repo "%s"' % (repo['name']))

			repository = self.__open_repo(repo)

			if repository == None:
				# a second shot
				repository = self.__open_repo(repo)

			if repository == None:
				continue

			# git fetch origin
			#repository.remotes.origin.fetch('+refs/heads/*:refs/remotes/origin/*')
			repository.remotes.origin.fetch()

			# git checkout
			repository.git.checkout(repo['branch'])

			# git log
			log = repository.git.log(self.__log_param)

			# split the log into lines
			commits = log.splitlines()

			print('- %d commit(s) found' % (len(commits)))

			for commit in commits:
				item = commit.split('\t')
				if len(item) != 6:
					continue

				# already found in other repo
				commit_hash = item[0]

				if commit_hash in hash_cache:
					continue

				hash_cache.append(commit_hash)

				author_email = item[1]
				author_date = item[2]
				committer_email = item[3]
				committer_date = item[4]
				subject = item[5]
				if repo['name'] == 'linux':
					status = 'upstreamed'
				else:
					status = 'accepted' # waiting next merge window

				user = self.get_user(email = author_email)

				if user == None:
					# should not happen
					user['name'] = 'John Doe'
					user['function'] = 'Dead man'

				self.__commits.append({'user_name': user['name'],
						       'user_function': user['function'],
						       'commit_hash': commit_hash,
						       'author_email': author_email,
						       'author_date': author_date,
						       'committer_email': committer_email,
						       'committer_date': committer_date,
						       'subject': subject,
						       'status': status,
						      })

		# sort the commits by date
		self.__commits.sort(key = useDateTime)

		return self.__commits

	def export_csv_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_csv_file(report_directory, self.__report_name, self.__csv_fields, self.__commits)

		return ret

	def export_excel_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_excel_file(report_directory, self.__report_name, self.__csv_fields, 'committer_date', self.__commits)

		return ret

class GithubCrawler(BaseCrawler):
	__csv_fields = ['user_name', 'user_function', 'repo_name', 'repo_url', 'number', 'state', 'title', 'user', 'created_at', 'updated_at', 'closed_at', 'merged_at', 'head', 'base', 'commits', 'additions', 'deletions', 'changed_files']
	__report_name = 'github-pulls'

	def __init__(self, cfg_path, auth):
		self.__repos = []
		self.__initialized = False

		# call parent's init
		super().__init__(cfg_path)

		if super().get_initialized() == False:
			return

		config = super().get_config()

		for key in config:
			if key.split(' ')[0] == 'github':
				if config[key]['disable'].lower() != 'false':
					continue

				self.__repos.append({'name': config[key]['name'],
						     'owner/repo': config[key]['owner/repo'],
						    })

		# save the github (username, token) pair
		self.__auth = auth

		self.__initialized = True

		return

	def get_pulls(self):
		# github REST API doc:
		# https://docs.github.com/en/rest

		def useDateTime(element):
			# 'created_at': '2019-06-11T09:10:12Z'
			return element['created_at']

		self.__pulls = []

		if self.__initialized == False:
			return self.__pulls

		# there is no filter for submitter
		usernames = []
		for user in self.get_users():
			usernames.append(user['github username'])

		for repo in self.__repos:
			print('query pulls from github repo "%s"' % (repo['name']))

			found = checked = 0

			# get the page one result
			url = 'https://api.github.com/repos/%s/pulls?state=all&per_page=100&direction=asc' % (repo['owner/repo'])

			while True:
				try:
					r = requests.get(url = url, auth = self.__auth)
				except requests.exceptions.RequestException as error:
					print('- github error: %s' % (error.message))
					break

				pulls = r.json()

				invalid_pulls = False

				for pull in pulls:
					if type(pull) is not dict:
						print('- fail to get pulls from repo')
						invalid_pulls = True
						break

					if pull['user']['login'] not in usernames:
						continue

					# one valid pull is found but don't know who's the submitter
					found += 1
					user = self.get_user(github_username = pull['user']['login'])

					if user == None:
						# should not happen
						user['name'] = 'John Doe'
						user['function'] = 'Dead man'

					detail = requests.get(url = pull['url'], auth = self.__auth).json()

					# check the response of 'GET /repos/{owner}/{repo}/pulls'
					# https://docs.github.com/en/rest/reference/pulls
					self.__pulls.append({'user_name': user['name'],
							     'user_function': user['function'],
							     'repo_name': repo['name'],
							     'repo_url': 'github.com/%s' % (repo['owner/repo']),
							     'number': pull['number'],
							     'state': pull['state'],
							     'title': pull['title'],
							     'user': pull['user']['login'],
							     'created_at': pull['created_at'],
							     'updated_at': pull['updated_at'],
							     'closed_at': pull['closed_at'],
							     'merged_at': pull['merged_at'],
							     'head': pull['head']['label'],
							     'base': pull['base']['label'],
							     'commits': detail['commits'],
							     'additions': detail['additions'],
							     'deletions': detail['deletions'],
							     'changed_files': detail['changed_files'],
							    })

				if invalid_pulls != False:
					# try next repo
					break

				checked += len(pulls)

				print('- %d pull(s) found / total %d pull(s) checked' % (found, checked))

				# read url to next page
				links = r.links
				if 'next' in links.keys():
					url = links['next']['url']
				else:
					break

		# sort the pulls by date
		self.__pulls.sort(key = useDateTime)

		return self.__pulls

	def export_csv_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_csv_file(report_directory, self.__report_name, self.__csv_fields, self.__pulls)

		return ret

	def export_excel_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_excel_file(report_directory, self.__report_name, self.__csv_fields, 'created_at', self.__pulls)

		return ret

class PatchworkCrawler(BaseCrawler):
	__csv_fields = ['user_name', 'user_function', 'repo_name', 'repo_url', 'project', 'date', 'name', 'state', 'submitter']
	__report_name = 'patchwork-patches'

	def __init__(self, cfg_path):
		self.__servers = []
		self.__initialized = False

		# call parent's init
		super().__init__(cfg_path)

		if super().get_initialized() == False:
			return

		config = super().get_config()

		for key in config:
			if key.split(' ')[0] == 'patchwork':
				if config[key]['disable'].lower() != 'false':
					continue

				self.__servers.append({'name': config[key]['name'],
						       'url': config[key]['url'],
						      })

		self.__initialized = True

		return

	def get_patches(self):
		# patchwork REST API doc:
		# https://patchwork.readthedocs.io/en/latest/api/rest/

		def useDateTime(element):
			# 'date': '2018-04-24T11:15:52'
			return element['date']

		self.__patches = []

		if self.__initialized == False:
			return self.__patches

		for server in self.__servers:
			print('query patches from patchwork server "%s"' % (server['name']))

			for user in self.get_users():
				for email in user['emails']:
					print('query patches for "%s <%s>"' % (user['name'], email))

					# get the page one result
					url = 'https://%s/api/1.2/patches?submitter=%s' % (server['url'], email)

					while True:
						try:
							r = requests.get(url = url)
						except requests.exceptions.RequestException as error:
							print('- patchwork error: %s' % (error.message))
							break

						patches = r.json()

						print('- %d patche(s) found' % (len(patches)))

						for patch in patches:
							# check the response of 'GET /api/1.2/patches/'
							# https://patchwork.readthedocs.io/en/latest/api/rest/schemas/v1.2/
							self.__patches.append({'user_name': user['name'],
									       'user_function': user['function'],
									       'repo_name': server['name'],
									       'repo_url': server['url'],
									       'project': patch['project']['name'],
									       'date': patch['date'],
									       'name': patch['name'],
									       'state': patch['state'],
									       'submitter': email,
									      })

						# read url to next page
						links = r.links
						if 'next' in links.keys():
							url = links['next']['url']
						else:
							break

		# sort the patches by date
		self.__patches.sort(key = useDateTime)

		return self.__patches

	def export_csv_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_csv_file(report_directory, self.__report_name, self.__csv_fields, self.__patches)

		return ret

	def export_excel_file(self, report_directory):
		if self.__initialized == False:
			return False

		ret = super().export_excel_file(report_directory, self.__report_name, self.__csv_fields, 'date', self.__patches)

		return ret
